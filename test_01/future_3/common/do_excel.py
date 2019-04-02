#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from future_3.common import project_path
from future_3.common.configrue_msg import ReadConfig

class DoExcel:
    '''该类完成测试数据的读取，测试数据的写回'''

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name  # excel 文件名
        self.sheet_name = sheet_name  # excel 表单名

    def read_excel(self,section):
        '''读取excel的信息'''
        case_id = ReadConfig(project_path.conf_path).get_data(section,'case_id')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        tel = self.get_tel()
        excel_list = []
        for i in range(2, sheet.max_row + 1):
            row_dict = {}
            row_dict['CaseId'] = sheet.cell(row=i, column=1).value
            row_dict['Module'] = sheet.cell(row=i, column=2).value
            row_dict['Description'] = sheet.cell(row=i, column=3).value
            row_dict['Method'] = sheet.cell(row=i, column=4).value
            row_dict['Url'] = sheet.cell(row=i, column=5).value
            if sheet.cell(i,6).value.find('tel')!=-1:#找到tel这个字符串
                row_dict['Params'] = sheet.cell(row=i, column=6).value.replace('tel',str(tel+1))
                self.write_excel(2,7,tel+1)
            else:
                row_dict['Params'] = sheet.cell(row=i, column=6).value
            row_dict['sql'] = sheet.cell(row=i, column=7).value
            row_dict['ExpectedResult'] = sheet.cell(row=i, column=8).value
            excel_list.append(row_dict)
        wb.close()
        result_data =[]
        if case_id =='all':#获取所有用例数据
            result_data = excel_list
        else:
            for i in case_id:
                result_data.append(excel_list[i-1])
        return result_data

    def get_tel(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        wb.close()
        return sheet.cell(2,7).value

    def write_excel(self, row, column, value):
        '''往excle里写入信息'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, column, value)
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    t = DoExcel(project_path.cases_path, 'loan')
    res = t.read_excel('LOAN')
    # res = t.get_tel()
    print(res)
    # print(type(res))
    # res = t.write_excel(2,8,1)
